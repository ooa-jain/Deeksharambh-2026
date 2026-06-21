"""
Deeksharambh 2026 — Flask + MongoDB Atlas
Production-ready for ~1000 concurrent users.
Gunicorn + eventlet: gunicorn -k eventlet -w 1 --threads 8 -b 0.0.0.0:5000 app:app
"""

import os, uuid, logging, json, secrets
from datetime import datetime, timezone, timedelta
from functools import wraps

from flask import (Flask, render_template, request, jsonify,
                   Response, redirect, url_for, session, make_response)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_caching import Cache
from pymongo import MongoClient, ASCENDING, DESCENDING
from pymongo.errors import PyMongoError
from bson import ObjectId
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import io
from dotenv import load_dotenv

load_dotenv()

# ─── App ───────────────────────────────────────────────────────────────────────
app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-env")
# 30 days so the "already submitted" state survives across visits
app.permanent_session_lifetime = timedelta(days=30)

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ─── Rate limiter ───────────────────────────────────────────────────────────────
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per minute"],
    storage_uri=os.environ.get("REDIS_URL", "memory://"),
)

# ─── Cache ─────────────────────────────────────────────────────────────────────
cache = Cache(app, config={
    "CACHE_TYPE": "SimpleCache",
    "CACHE_DEFAULT_TIMEOUT": 60,
})

# ─── MongoDB ────────────────────────────────────────────────────────────────────
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017")
DB_NAME   = os.environ.get("DB_NAME",   "deeksharambh2026")

client = MongoClient(
    MONGO_URI,
    maxPoolSize=50,
    minPoolSize=5,
    serverSelectionTimeoutMS=5000,
    connectTimeoutMS=5000,
)
db        = client[DB_NAME]
responses = db["responses"]

# Indexes
responses.create_index([("timestamp", DESCENDING)])
responses.create_index([("school",    ASCENDING)])
responses.create_index([("ugpg",      ASCENDING)])
responses.create_index([("submitted", ASCENDING)])

# ─── Admin credentials ─────────────────────────────────────────────────────────
ADMIN_USER = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASS = os.environ.get("ADMIN_PASSWORD", "adminjain")


# ─── Helpers ───────────────────────────────────────────────────────────────────
def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


def json_safe(obj):
    if isinstance(obj, list):
        return [json_safe(o) for o in obj]
    if isinstance(obj, dict):
        return {k: json_safe(v) for k, v in obj.items() if k != "_id"}
    if isinstance(obj, ObjectId):
        return str(obj)
    if isinstance(obj, datetime):
        return obj.isoformat()
    return obj


# ─── Emoji → numeric map (used in Excel export for emoji-scale questions) ──────
EMOJI_TO_NUM = {
    # Q5 / Q6 transition ease
    "Very hard": 1, "Tough": 2, "Okay": 3, "Smooth": 4, "Super easy": 5,
    # Q10 binge
    "Absolutely!": 4, "Maybe": 3, "Not sure": 2, "Probably not": 1,
    # Q14 hands-on
    "All sitting, no doing": 1, "All sitting": 1,
    "Mostly passive": 2, "Some activities": 3,
    "Quite hands-on": 4, "Fully interactive!": 5, "Fully interactive": 5,
    # Q15 engagement
    "Sleep Mode": 1, "Interesting": 3,
    "Super Engaging": 4, "Couldn't Stop": 5,
    # Q20 NEP before
    "No idea": 1, "Heard of it": 2, "Basic idea": 3, "I knew well": 4,
    # Q22 digital confidence
    "Totally lost": 1, "Struggling": 2, "Getting there": 3,
    "Fairly confident": 4, "I got this!": 5,
    # Q25 first week
    "A rollercoaster": 3, "A blur": 2, "A celebration": 5,
    "Study mode": 3, "Fresh start": 4, "Survive mode": 2,
    # Q35 overall experience
    "Not great": 1, "Could be better": 2, "It was okay": 3,
    "Pretty good!": 4, "Absolutely loved it": 5,
}


def normalise_submission(data: dict) -> dict:
    """
    The new form stores transition ease as q5 (emoji text label).
    Older form used q5 as a numeric int. Both are stored as-is;
    the admin dashboard handles both via EMOJI_TO_NUM mapping.
    Also ensures q1, q4, q36, q40 are always flat lists.
    """
    # Flatten any nested list-of-lists (shouldn't happen but be safe)
    for key in ("q1", "q4", "q36", "q40", "q5a", "q5b",
                "q7", "q9", "q11", "q12", "q17", "q19",
                "q23a", "q23b", "q26", "q27", "q28",
                "q30", "q31", "q33", "q37", "q38", "q39"):
        v = data.get(key)
        if isinstance(v, list):
            flat = []
            for item in v:
                if isinstance(item, list):
                    flat.extend(item)
                elif item:
                    flat.append(item)
            data[key] = flat

    # Convert emoji-scale text values to numeric for easier Excel/stats
    for key in ("q5", "q10", "q14", "q15", "q20", "q22", "q25", "q35"):
        v = data.get(key)
        if isinstance(v, str) and v in EMOJI_TO_NUM:
            data[key + "_label"] = v          # keep original label
            data[key] = EMOJI_TO_NUM[v]       # store numeric too

    return data


# ─── Public routes ─────────────────────────────────────────────────────────────
@app.route("/")
def index():
    """Serve the student survey form.

    If the current browser has already submitted the survey (tracked by
    a server-side cookie session), redirect them straight to /result.
    """
    if session.get("survey_submitted"):
        return redirect(url_for("user_result"))
    return render_template("form.html")


@app.route("/result")
def user_result():
    """Display the thank-you / response snapshot for the current user.
    The user is identified by the server-side session `survey_submitted`
    flag plus the `response_id` we stored when they submitted.
    """
    if not session.get("survey_submitted"):
        return redirect(url_for("index"))

    response_id = session.get("response_id")
    doc = None
    if response_id:
        try:
            doc = responses.find_one({"id": response_id})
        except PyMongoError as e:
            log.error("user_result fetch error: %s", e)

    # Derive fields the result template needs
    avatar_text  = None
    submitted_at = None
    if doc:
        q41 = doc.get("q41")
        if isinstance(q41, list) and q41:
            avatar_text = q41[0]
        elif isinstance(q41, str) and q41:
            avatar_text = q41

        created = doc.get("created_at")
        if isinstance(created, datetime):
            submitted_at = created.strftime("%d %b %Y, %H:%M")
        elif created:
            submitted_at = str(created)

    return render_template(
        "result.html",
        response=doc,
        response_id=response_id,
        avatar_text=avatar_text,
        submitted_at=submitted_at,
    )


@limiter.limit("5 per minute")
@app.route("/api/submit", methods=["POST"])
def submit():
    try:
        data = request.get_json(force=True, silent=True) or {}
        if not data:
            return jsonify({"ok": False, "error": "Empty payload"}), 400

        data.pop("_id", None)
        data = normalise_submission(data)
        data["submitted"]  = True
        data["ip"]         = request.remote_addr
        data["user_agent"] = request.headers.get("User-Agent", "")[:200]
        data["created_at"] = datetime.now(timezone.utc)

        if not data.get("id"):
            data["id"] = str(uuid.uuid4())[:8]

        responses.insert_one(data)
        cache.clear()   # invalidate dashboard cache on new submission

        # ── Mark the user's session as having submitted ────────────────
        # This lets / and /result know on the next page load that this
        # browser already took the survey and should skip straight to
        # their result snapshot.
        session.permanent = True
        session["survey_submitted"] = True
        session["response_id"]      = data["id"]

        return jsonify({"ok": True, "id": data["id"]}), 201

    except PyMongoError as e:
        log.error("MongoDB error on submit: %s", e)
        return jsonify({"ok": False, "error": "DB error"}), 500
    except Exception as e:
        log.exception("Unexpected error on submit")
        return jsonify({"ok": False, "error": str(e)}), 500


# ─── User: clear their session (allow retake) ─────────────────────────────────
@app.route("/api/reset-session", methods=["POST"])
def reset_session():
    """Clears the server-side session flag so the user can take the
    survey again (useful for demo / testing)."""
    session.pop("survey_submitted", None)
    session.pop("response_id", None)
    return jsonify({"ok": True})


# ─── Admin auth ────────────────────────────────────────────────────────────────
@app.route("/admin/login", methods=["GET", "POST"])
@limiter.limit("10 per minute")
def admin_login():
    error = None
    if request.method == "POST":
        if (request.form.get("username") == ADMIN_USER and
                request.form.get("password") == ADMIN_PASS):
            session["admin_logged_in"] = True
            return redirect(url_for("admin_dashboard"))
        error = "Invalid credentials"
    return render_template("admin_login.html", error=error)


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


# ─── Admin dashboard ───────────────────────────────────────────────────────────
@app.route("/admin")
@app.route("/admin/dashboard")
@require_admin
def admin_dashboard():
    return render_template("admin.html")


# ─── Admin: delete selected ────────────────────────────────────────────────────
@app.route("/admin/api/delete_selected", methods=["POST"])
@require_admin
def delete_selected():
    try:
        data = request.get_json(force=True, silent=True) or {}
        ids  = data.get("ids", [])
        if not ids:
            return jsonify({"ok": False, "error": "No IDs provided"}), 400
        cache.clear()
        result = responses.delete_many({"id": {"$in": ids}})
        return jsonify({"ok": True, "deleted_count": result.deleted_count})
    except PyMongoError as e:
        log.error("delete_selected error: %s", e)
        return jsonify({"ok": False, "error": "DB error"}), 500


# ─── Admin: delete all ─────────────────────────────────────────────────────────
@app.route("/admin/api/delete_all", methods=["POST"])
@require_admin
def delete_all():
    try:
        cache.clear()
        result = responses.delete_many({})
        return jsonify({"ok": True, "deleted_count": result.deleted_count})
    except PyMongoError as e:
        log.error("delete_all error: %s", e)
        return jsonify({"ok": False, "error": "DB error"}), 500


# ─── Admin: stats ──────────────────────────────────────────────────────────────
def make_stats_cache_key():
    school = request.args.get("school", "all")
    ugpg   = request.args.get("ugpg",   "all")
    return f"admin_stats_{school}_{ugpg}"


@app.route("/admin/api/stats")
@require_admin
@cache.cached(timeout=60, make_cache_key=make_stats_cache_key)
def admin_stats():
    school = request.args.get("school", "all")
    ugpg   = request.args.get("ugpg",   "all")

    filt = {"submitted": True}
    if school != "all":
        filt["school"] = school
    if ugpg != "all":
        filt["ugpg"] = {"$regex": ugpg, "$options": "i"}

    try:
        docs = list(responses.find(filt, {"_id": 0, "ip": 0, "user_agent": 0}))
        return jsonify({"ok": True, "count": len(docs), "data": json_safe(docs)})
    except PyMongoError as e:
        log.error("admin_stats error: %s", e)
        return jsonify({"ok": False, "error": "DB error"}), 500


@app.route("/admin/api/schools")
@require_admin
def admin_schools():
    try:
        schools = responses.distinct("school", {"submitted": True})
        return jsonify({"ok": True, "schools": sorted(s for s in schools if s)})
    except PyMongoError:
        return jsonify({"ok": False, "error": "DB error"}), 500


# ─── Admin: Excel export ───────────────────────────────────────────────────────
@app.route("/admin/api/export")
@require_admin
def admin_export():
    school = request.args.get("school", "all")
    ugpg   = request.args.get("ugpg",   "all")

    filt = {"submitted": True}
    if school != "all":
        filt["school"] = school
    if ugpg != "all":
        filt["ugpg"] = {"$regex": ugpg, "$options": "i"}

    try:
        docs = list(responses.find(filt, {"_id": 0, "ip": 0, "user_agent": 0}))
    except PyMongoError:
        return "Database error", 500

    wb  = _build_excel(docs)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"Deeksharambh2026_Responses_{ts}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _build_excel(docs: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    NAVY  = "0D2147"; GOLD  = "E6B324"; GREEN = "1D9E75"
    AMBER = "BA7517"; RED   = "D85A30"; WHITE = "FFFFFF"
    LGREY = "F5F4FB"; MGREY = "EBEBF5"

    hdr_fill  = PatternFill("solid", fgColor=NAVY)
    hdr_font  = Font(bold=True, color=WHITE, size=10)
    gold_fill = PatternFill("solid", fgColor=GOLD)
    gold_font = Font(bold=True, color=NAVY, size=10)
    alt_fill  = PatternFill("solid", fgColor=LGREY)
    even_fill = PatternFill("solid", fgColor=WHITE)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCDD")
    thin_b    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Raw Responses ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Raw Responses"

    SCALAR_COLS = [
        ("id",          "Response ID",               12),
        ("created_at",  "Submitted At",              20),
        ("school",      "School",                    32),
        ("program",     "Programme",                 28),
        ("ugpg",        "UG / PG",                   10),
        ("q2",          "Q2 Overall Vibe (1-10)",    14),
        ("q3",          "Q3 Felt Welcomed",           22),
        ("q5",          "Q5 Transition Ease (1-5)",  18),
        ("q5_label",    "Q5 Transition Label",       18),
        ("q10",         "Q10 Footsteps Binge",       18),
        ("q16",         "Q16 Bridge Confidence(1-5)",18),
        ("q18",         "Q18 Ready for Classes",     22),
        ("q20",         "Q20 NEP Awareness Before",  18),
        ("q21",         "Q21 NEP Clarity After",     22),
        ("q22",         "Q22 Digital Confidence",    18),
        ("q29",         "Q29 Belonging (1-10)",      16),
        ("q32",         "Q32 Success Confidence(1-10)",18),
        ("q34",         "Q34 NPS (0-10)",            14),
        ("q35",         "Q35 Overall Experience",    22),
        ("q41",         "Q41 Avatar",                18),
    ]

    MULTI_COLS = [
        ("q1",   "Q1 Vibe Words"),
        ("q4",   "Q4 Week Feelings"),
        ("q5a",  "Q5a Spark Sessions"),
        ("q5b",  "Q5b Missed Sessions"),
        ("q7",   "Q7 Onboarding Challenges"),
        ("q9",   "Q9 Footsteps Content"),
        ("q11",  "Q11 High-Impact Sessions"),
        ("q12",  "Q12 Sessions to Improve"),
        ("q17",  "Q17 Bridge Areas"),
        ("q19",  "Q19 Helpful Aspects"),
        ("q23a", "Q23a Expectations Before"),
        ("q23b", "Q23b Expectations After"),
        ("q26",  "Q26 Biggest Surprise"),
        ("q27",  "Q27 Smile Moments"),
        ("q28",  "Q28 Stress Points"),
        ("q30",  "Q30 Friends Made"),
        ("q31",  "Q31 Knows Who to Contact"),
        ("q33",  "Q33 Academic Expectations"),
        ("q36",  "Q36 NPS Reason"),
        ("q37",  "Q37 Keep"),
        ("q38",  "Q38 Stop"),
        ("q39",  "Q39 Introduce"),
        ("q40",  "Q40 Final Feeling"),
    ]

    all_cols = SCALAR_COLS + [(k, lbl, 38) for k, lbl in MULTI_COLS]

    for ci, (key, label, width) in enumerate(all_cols, 1):
        cell = ws1.cell(row=1, column=ci, value=label)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = thin_b
        ws1.column_dimensions[get_column_letter(ci)].width = width

    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "A2"

    for ri, doc in enumerate(docs, 2):
        fill = alt_fill if ri % 2 == 0 else even_fill
        for ci, (key, _, _w) in enumerate(all_cols, 1):
            val = doc.get(key, "")
            if isinstance(val, list):
                val = " | ".join(str(v) for v in val if v)
            elif isinstance(val, dict):
                val = json.dumps(val)
            elif isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d %H:%M")
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.alignment = left; cell.border = thin_b

    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: Summary Dashboard ─────────────────────────────────────────
    ws2 = wb.create_sheet("Summary Dashboard")
    n = len(docs)

    def safe_avg(key):
        vals = [float(d.get(key, 0) or 0) for d in docs
                if str(d.get(key, "")).replace(".", "").lstrip("-").isdigit()]
        return round(sum(vals) / len(vals), 1) if vals else 0

    def nps_score():
        sc = [int(d.get("q34", 0)) for d in docs
              if str(d.get("q34", "")).isdigit()]
        le = len(sc) or 1
        return round((sum(1 for s in sc if s >= 9) / le -
                      sum(1 for s in sc if s < 7)  / le) * 100)

    def kpi(ws, row, col, label, value, color=NAVY):
        lc = ws.cell(row=row,   column=col, value=label)
        vc = ws.cell(row=row+1, column=col, value=value)
        lc.fill = PatternFill("solid", fgColor=color)
        lc.font = Font(bold=True, color=WHITE, size=9)
        lc.alignment = center
        vc.font = Font(bold=True, color=color, size=16)
        vc.alignment = center
        for c in [lc, vc]:
            c.border = thin_b
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws2.merge_cells("A1:H1")
    t = ws2["A1"]
    t.value = "DEEKSHARAMBH 2026 — SUMMARY DASHBOARD"
    t.font = Font(bold=True, color=WHITE, size=14)
    t.fill = hdr_fill; t.alignment = center
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells("A2:H2")
    s = ws2["A2"]
    s.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total Responses: {n}"
    s.font  = Font(color=NAVY, size=10, italic=True)
    s.fill  = PatternFill("solid", fgColor=MGREY)
    s.alignment = center
    ws2.row_dimensions[2].height = 18

    kpi(ws2, 4, 1, "Total Responses",         n,               NAVY)
    kpi(ws2, 4, 2, "Avg Vibe (Q2)",           safe_avg("q2"),  GREEN)
    kpi(ws2, 4, 3, "Avg Belonging (Q29)",     safe_avg("q29"), RED)
    kpi(ws2, 4, 4, "NPS Score",               nps_score(),     AMBER)
    kpi(ws2, 4, 5, "Transition Ease (Q5)",    safe_avg("q5"),  GREEN)
    kpi(ws2, 4, 6, "Bridge Confidence (Q16)", safe_avg("q16"), NAVY)
    kpi(ws2, 4, 7, "Success Conf. (Q32)",     safe_avg("q32"), GREEN)
    kpi(ws2, 4, 8, "Avg Experience (Q35)",    safe_avg("q35"), RED)

    def top_items(key, n_items=5):
        freq = {}
        for d in docs:
            v = d.get(key, [])
            if isinstance(v, list):
                for item in v:
                    if item: freq[item] = freq.get(item, 0) + 1
            elif v:
                freq[str(v)] = freq.get(str(v), 0) + 1
        return sorted(freq.items(), key=lambda x: -x[1])[:n_items]

    sections = [
        (7,  "TOP VIBE WORDS (Q1)",          "q1"),
        (14, "TOP STRESS POINTS (Q28)",       "q28"),
        (21, "TOP SMILE MOMENTS (Q27)",       "q27"),
        (28, "TOP SESSIONS TO KEEP (Q37)",    "q37"),
        (35, "TOP ADD NEXT YEAR (Q39)",       "q39"),
        (42, "TOP SESSIONS TO IMPROVE (Q12)","q12"),
    ]

    for start_row, title, key in sections:
        ws2.merge_cells(f"A{start_row}:H{start_row}")
        h = ws2[f"A{start_row}"]
        h.value = title; h.font = gold_font; h.fill = gold_fill
        h.alignment = center
        ws2.row_dimensions[start_row].height = 22
        ws2.cell(row=start_row+1, column=1, value="Item").font = Font(bold=True, size=9)
        ws2.cell(row=start_row+1, column=5, value="Count").font = Font(bold=True, size=9)
        ws2.cell(row=start_row+1, column=6, value="%").font = Font(bold=True, size=9)
        for idx, (item, cnt) in enumerate(top_items(key), start_row+2):
            ws2.cell(row=idx, column=1, value=item).alignment = left
            ws2.merge_cells(f"A{idx}:D{idx}")
            ws2.cell(row=idx, column=5, value=cnt)
            ws2.cell(row=idx, column=6, value=f"{round(cnt/n*100)}%" if n else "0%")

    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ── Sheet 3: School Breakdown ──────────────────────────────────────────
    ws3 = wb.create_sheet("By School")
    ws3.merge_cells("A1:G1")
    h1 = ws3["A1"]
    h1.value = "RESPONSES BY SCHOOL"
    h1.font  = Font(bold=True, color=WHITE, size=12)
    h1.fill  = hdr_fill; h1.alignment = center
    ws3.row_dimensions[1].height = 28

    headers3 = ["School", "Responses", "% Share",
                 "Avg Vibe", "Avg Belong", "NPS", "Avg Transition"]
    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.alignment = center; cell.border = thin_b

    schools_map = {}
    for d in docs:
        sc = d.get("school", "Unknown") or "Unknown"
        schools_map.setdefault(sc, []).append(d)

    ri = 3
    for sc, sdocs in sorted(schools_map.items(), key=lambda x: -len(x[1])):
        sn = len(sdocs)
        sc_nps = [int(d.get("q34", 0)) for d in sdocs
                  if str(d.get("q34", "")).isdigit()]
        sl = len(sc_nps) or 1
        sc_nps_val = round((sum(1 for s in sc_nps if s >= 9) / sl -
                            sum(1 for s in sc_nps if s < 7)  / sl) * 100)

        # avg transition for this school (numeric q5)
        sc_trans = [float(d.get("q5", 0) or 0) for d in sdocs
                    if str(d.get("q5", "")).replace(".", "").isdigit()]
        sc_trans_avg = round(sum(sc_trans)/len(sc_trans), 1) if sc_trans else 0

        sc_vibe = [float(d.get("q2", 0) or 0) for d in sdocs
                   if str(d.get("q2", "")).replace(".", "").isdigit()]
        sc_vibe_avg = round(sum(sc_vibe)/len(sc_vibe), 1) if sc_vibe else 0

        sc_bel = [float(d.get("q29", 0) or 0) for d in sdocs
                  if str(d.get("q29", "")).replace(".", "").isdigit()]
        sc_bel_avg = round(sum(sc_bel)/len(sc_bel), 1) if sc_bel else 0

        row_fill = alt_fill if ri % 2 == 0 else even_fill
        vals = [sc, sn, f"{round(sn/n*100)}%", sc_vibe_avg, sc_bel_avg,
                sc_nps_val, sc_trans_avg]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.fill = row_fill
            cell.alignment = center if ci > 1 else left
            cell.border = thin_b
        ri += 1

    ws3.column_dimensions["A"].width = 38
    for col in "BCDEFG":
        ws3.column_dimensions[col].width = 16
    ws3.freeze_panes = "A3"
    ws3.auto_filter.ref = f"A2:{get_column_letter(len(headers3))}2"

    # ── Sheet 4: NPS Distribution ──────────────────────────────────────────
    ws4 = wb.create_sheet("NPS Distribution")
    ws4.merge_cells("A1:C1")
    h1 = ws4["A1"]
    h1.value = "NPS SCORE DISTRIBUTION"
    h1.font = Font(bold=True, color=WHITE, size=12)
    h1.fill = hdr_fill; h1.alignment = center

    for ci, h in enumerate(["Score", "Count", "%"], 1):
        c = ws4.cell(row=2, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.alignment = center

    nps_freq = {i: 0 for i in range(11)}
    for d in docs:
        v = d.get("q34")
        if v is not None and str(v).isdigit():
            nps_freq[int(v)] = nps_freq.get(int(v), 0) + 1

    for ri2, (score, cnt) in enumerate(sorted(nps_freq.items()), 3):
        cat = GREEN if score >= 9 else (AMBER if score >= 7 else RED)
        ws4.cell(row=ri2, column=1, value=score).alignment = center
        ws4.cell(row=ri2, column=2, value=cnt).alignment   = center
        pct_cell = ws4.cell(row=ri2, column=3,
                             value=f"{round(cnt/n*100)}%" if n else "0%")
        pct_cell.alignment = center
        pct_cell.font = Font(color=cat, bold=True)

    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 12

    return wb


# ─── Health-check ──────────────────────────────────────────────────────────────
@app.route("/health")
def health():
    try:
        client.admin.command("ping")
        return jsonify({"status": "ok", "db": "connected"}), 200
    except Exception as e:
        return jsonify({"status": "error", "detail": str(e)}), 503


# ─── Entry-point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1",
            host="0.0.0.0", port=5000, threaded=True)
